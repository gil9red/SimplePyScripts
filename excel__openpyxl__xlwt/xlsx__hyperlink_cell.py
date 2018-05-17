#!/usr/bin/env python3
# -*- coding: utf-8 -*-

__author__ = 'ipetrash'


from urllib.parse import quote
import openpyxl


columns = ['Language', 'Text', 'Hyperlink']
rows = [
    ['python', 'excel'],
    ['java', 'excel'],
    ['c#', 'excel'],
]

wb = openpyxl.Workbook()
ws = wb.get_active_sheet()

for i, value in enumerate(columns, 1):
    ws.cell(row=1, column=i).value = value

for i, row in enumerate(rows, 2):
    lang, text = row

    ws.cell(row=i, column=1).value = lang
    ws.cell(row=i, column=2).value = text

    ws.cell(row=i, column=3).hyperlink = "https://stackoverflow.com/search?q=" + quote(lang + ' ' + text)
    ws.cell(row=i, column=3).value = "StackOverflow"


from openpyxl.writer.excel import save_workbook
save_workbook(wb, 'excel.xlsx')
