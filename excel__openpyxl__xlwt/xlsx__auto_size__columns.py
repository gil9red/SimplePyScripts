#!/usr/bin/env python3
# -*- coding: utf-8 -*-

__author__ = 'ipetrash'


import openpyxl


def set_column_size(worksheet):
    dims = dict()

    for row in worksheet.rows:
        for cell in row:
            if not cell.value:
                continue

            dims[cell.column] = max(dims.get(cell.column, 0), len(str(cell.value)))

    for col, value in dims.items():
        worksheet.column_dimensions[col].width = value * 1.15  # Append 15%


def fill_sheet(ws):
    columns = ['Language', 'Text']
    rows = [
        ['python', '-' * 10],
        ['java', 'j' * 10],
        ['c#', '*' * 5],
        ['c++', '0' * 50],
    ]

    for i, value in enumerate(columns, 1):
        ws.cell(row=1, column=i).value = value

    for i, row in enumerate(rows, 2):
        for j, value in enumerate(row, 1):
            cell = ws.cell(row=i, column=j)
            cell.value = value


wb = openpyxl.Workbook()
ws = wb.get_active_sheet()

# Sheet 1
fill_sheet(ws)

# Sheet 2
ws = wb.create_sheet('auto')
fill_sheet(ws)
set_column_size(ws)

from openpyxl.writer.excel import save_workbook
save_workbook(wb, 'excel.xlsx')
