#!/usr/bin/env python3
# -*- coding: utf-8 -*-

__author__ = 'ipetrash'


import openpyxl


columns = ['Language', 'Text', 'Total']
rows = [
    ['python', 1],
    ['java', 2],
    ['c#', 3],
]

wb = openpyxl.Workbook()
ws = wb.get_active_sheet()

for i, value in enumerate(columns, 1):
    ws.cell(row=1, column=i).value = value

for i, row in enumerate(rows, 2):
    for j, value in enumerate(row, 1):
        cell = ws.cell(row=i, column=j)
        cell.value = value

ws.merge_cells('C2:C4')

# Total:
ws['C2'].value = '=SUM(B2:B4)'

from openpyxl.writer.excel import save_workbook
save_workbook(wb, 'excel.xlsx')
