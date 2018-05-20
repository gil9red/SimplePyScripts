#!/usr/bin/env python3
# -*- coding: utf-8 -*-

__author__ = 'ipetrash'


import openpyxl
from openpyxl.utils.cell import _get_column_letter


def set_row_column_size(ws):
    for i in range(1, 500):
        row = ws.row_dimensions[i]
        row.height = 15

        for j in range(1, 500):
            j = _get_column_letter(j)

            col = ws.column_dimensions[j]
            col.width = 3.5


wb = openpyxl.Workbook()
ws = wb.get_active_sheet()

# Sheet 2
ws = wb.create_sheet()
set_row_column_size(ws)

from openpyxl.writer.excel import save_workbook
save_workbook(wb, 'excel.xlsx')
