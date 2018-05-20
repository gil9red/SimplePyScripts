#!/usr/bin/env python3
# -*- coding: utf-8 -*-

__author__ = 'ipetrash'


import openpyxl

wb = openpyxl.Workbook()

ws = wb.get_active_sheet()
ws.title = '20'
ws.sheet_view.zoomScale = 20

ws = wb.create_sheet('50')
ws.sheet_view.zoomScale = 50

ws = wb.create_sheet('100')

ws = wb.create_sheet('200')
ws.sheet_view.zoomScale = 200

ws = wb.create_sheet('500')
ws.sheet_view.zoomScale = 500

from openpyxl.writer.excel import save_workbook
save_workbook(wb, 'excel.xlsx')
