#!/usr/bin/env python3
# -*- coding: utf-8 -*-

__author__ = 'ipetrash'


# pip install openpyxl
import openpyxl


columns = ['Name', 'Age', 'Course']
rows = [
    ['Vasya', '16', 1],
    ['Anya', '17', 2],
    ['Inna', '16', 1],
]


wb = openpyxl.Workbook()

# Remove default sheet
for sheet_name in wb.sheetnames:
    sheet = wb[sheet_name]
    wb.remove(sheet)

ws = wb.create_sheet('Students')

for i, value in enumerate(columns, 1):
    ws.cell(row=1, column=i).value = value

for i, row in enumerate(rows, 2):
    for j, value in enumerate(row, 1):
        ws.cell(row=i, column=j).value = value

wb.save('excel.xlsx')
