#!/usr/bin/env python3
# -*- coding: utf-8 -*-

__author__ = 'ipetrash'


import openpyxl


columns = ['Language', 'Text']
rows = [
    ['python', 1],
    ['java', 2],
    ['c#', 3],
    ['Total:', 0]
]

wb = openpyxl.Workbook()
ws = wb.active

for i, value in enumerate(columns, 1):
    ws.cell(row=1, column=i).value = value

for i, row in enumerate(rows, 2):
    for j, value in enumerate(row, 1):
        cell = ws.cell(row=i, column=j)
        cell.value = value

# Total:
ws.cell(row=5, column=2).value = '=SUM(B2:B4)'

wb.save('excel.xlsx')
