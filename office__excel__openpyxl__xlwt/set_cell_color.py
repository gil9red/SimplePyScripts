#!/usr/bin/env python3
# -*- coding: utf-8 -*-

__author__ = 'ipetrash'


import openpyxl
from openpyxl.styles import PatternFill


wb = openpyxl.Workbook()
ws = wb.active

ws.cell(row=1, column=1).fill = PatternFill(fgColor='FF0000', fill_type="solid")
ws.cell(row=2, column=1).fill = PatternFill(fgColor='00FF00', fill_type="solid")
ws['A3'].fill = PatternFill(fgColor='0000FF', fill_type="solid")

wb.save('excel.xlsx')
