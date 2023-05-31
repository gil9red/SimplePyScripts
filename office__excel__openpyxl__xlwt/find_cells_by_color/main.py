#!/usr/bin/env python3
# -*- coding: utf-8 -*-

__author__ = "ipetrash"


# SOURCE: https://ru.stackoverflow.com/a/845335/201445


from openpyxl.worksheet.worksheet import Worksheet


def find_cells_by_color(ws: Worksheet, color: str = "00000000") -> dict:
    ret = dict()
    for row in ws.iter_rows():
        for cell in row:
            if cell.fill.fgColor.value == color:
                ret[cell.coordinate] = cell.value

    return ret


if __name__ == "__main__":
    from openpyxl import load_workbook

    wb = load_workbook("excel.xlsx")
    ws = wb.active

    print("background colors for ALL cells:\n")
    for row in ws.iter_rows():
        for cell in row:
            print(f"[{cell.coordinate}]: {cell.fill.fgColor.value}", end=" ")

        print()

    print()

    cells = find_cells_by_color(ws, color="FFFFFF00")
    print(f"given color has been found in the following cells: {cells}")
