#!/usr/bin/env python3
# -*- coding: utf-8 -*-

__author__ = "ipetrash"


import openpyxl
from openpyxl.styles import PatternFill

# pip install Pillow
from PIL import Image


# SOURCE: https://github.com/gil9red/SimplePyScripts/blob/8b69f96d0bd87c10f3d0779ff46b02558c4cca84/excel__openpyxl__xlwt/xlsx__set_row_column_size.py
def set_row_column_size(ws) -> None:
    # SOURCE: excel__openpyxl__xlwt\excel\xl\worksheets\sheet1.xml
    # <sheetFormatPr defaultColWidth="1.77734375" defaultRowHeight="6.6" customHeight="1" x14ac:dyDescent="0.3"/>
    ws.sheet_format.defaultColWidth = 1.77734375
    ws.sheet_format.defaultRowHeight = 9.0
    ws.sheet_format.customHeight = 1


# SOURCE: https://github.com/gil9red/SimplePyScripts/blob/72dc18705fa99b3346e836938dc9410d6b4a2a21/pil_pillow__examples/for_cycle__pixels__array/main.py
def get_pixel_array(img, rgb_hex=False):
    width, height = img.size

    pixels = []

    for y in range(height):
        row = []
        pixels.append(row)

        for x in range(width):
            r, g, b = img.getpixel((x, y))

            if rgb_hex:
                value = f"{r:02X}{g:02X}{b:02X}"
                row.append(value)
            else:
                row.append((r, g, b))

    return pixels


def draw_image(ws, img) -> None:
    img = img.convert("RGB")

    # Resize
    img.thumbnail((250, 250))

    pixels = get_pixel_array(img, rgb_hex=True)

    for i in range(len(pixels)):
        for j in range(len(pixels[0])):
            cell = ws.cell(row=i + 1, column=j + 1)
            cell.fill = PatternFill(fgColor=pixels[i][j], fill_type="solid")


if __name__ == "__main__":
    wb = openpyxl.Workbook()
    ws = wb.active

    # Масштаб 10%
    ws.sheet_view.zoomScale = 10

    set_row_column_size(ws)

    file_name = "input.jpg"
    img = Image.open(file_name)
    ws.title = file_name
    draw_image(ws, img)

    wb.save("excel.xlsx")
