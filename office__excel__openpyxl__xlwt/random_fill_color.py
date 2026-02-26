#!/usr/bin/env python3
# -*- coding: utf-8 -*-

__author__ = "ipetrash"


import random

import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.worksheet.worksheet import Worksheet


# SOURCE: https://github.com/gil9red/SimplePyScripts/blob/8b69f96d0bd87c10f3d0779ff46b02558c4cca84/excel__openpyxl__xlwt/xlsx__set_row_column_size.py
def set_row_column_size(ws: Worksheet) -> None:
    # SOURCE: excel__openpyxl__xlwt\excel\xl\worksheets\sheet1.xml
    # <sheetFormatPr defaultColWidth="1.77734375" defaultRowHeight="6.6" customHeight="1" x14ac:dyDescent="0.3"/>
    ws.sheet_format.defaultColWidth = 1.77734375
    ws.sheet_format.defaultRowHeight = 9.0
    ws.sheet_format.customHeight = 1


# SOURCE: https://github.com/gil9red/SimplePyScripts/blob/1c2274804f52004b07601ece535ec74b3f82aa8d/get_random_hex_color.py
def get_random_hex_color() -> str:
    return "".join(random.choices("0123456789ABCDEF", k=6))


def set_fill_color(ws: Worksheet) -> None:
    size = 250

    for i in range(1, size + 1):
        for j in range(1, size + 1):
            cell = ws.cell(row=i, column=j)
            cell.fill = PatternFill(fgColor=get_random_hex_color(), fill_type="solid")


wb = openpyxl.Workbook()
ws = wb.active

# Масштаб 10%
ws.sheet_view.zoomScale = 10

set_row_column_size(ws)
set_fill_color(ws)

wb.save("excel.xlsx")
