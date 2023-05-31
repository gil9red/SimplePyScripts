#!/usr/bin/env python3
# -*- coding: utf-8 -*-

__author__ = "ipetrash"


import openpyxl


path = "../get_sheets/excel.xlsx"

wb = openpyxl.load_workbook(path)
sheet = wb.active

for row in range(1, sheet.max_row + 1):
    for col in range(1, sheet.max_column + 1):
        cell = sheet.cell(row=row, column=col)
        print(cell.value, end=" | ")
    print()
"""
Language | Text | 
python | ---------- | 
java | jjjjjjjjjj | 
c# | ***** | 
c++ | 00000000000000000000000000000000000000000000000000 | 
"""
