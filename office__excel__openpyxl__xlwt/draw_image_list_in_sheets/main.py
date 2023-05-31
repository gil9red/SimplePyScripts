#!/usr/bin/env python3
# -*- coding: utf-8 -*-

__author__ = "ipetrash"


import glob
import os
import sys

import openpyxl

# pip install Pillow
from PIL import Image

# SOURCE: https://github.com/gil9red/SimplePyScripts/blob/2c4391214de936260926a47440f2bad2d6fb90da/excel__openpyxl__xlwt/draw_image_in_sheet
sys.path.append("..")
from draw_image_in_sheet.main import set_row_column_size, get_pixel_array, draw_image


if __name__ == "__main__":
    wb = openpyxl.Workbook()

    # Remove default sheet
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        wb.remove(sheet)

    # Append images
    for file_name in glob.glob("images/*.jpg") + glob.glob("images/*.png"):
        title = os.path.basename(file_name)
        ws = wb.create_sheet(title)

        # Масштаб 10%
        ws.sheet_view.zoomScale = 10

        set_row_column_size(ws)

        img = Image.open(file_name)
        draw_image(ws, img)

    wb.save("excel.xlsx")
