#!/usr/bin/env python3
# -*- coding: utf-8 -*-

__author__ = 'ipetrash'


import openpyxl
from openpyxl.worksheet.worksheet import Worksheet


def set_column_size(ws: Worksheet):
    dims: dict[str, int] = dict()

    for row in ws.rows:
        for cell in row:
            if not cell.value:
                continue

            column = cell.column_letter
            dims[column] = max(dims.get(column, 0), len(str(cell.value)))

    for col, value in dims.items():
        ws.column_dimensions[col].width = value * 1.15  # Append 15%


def fill_sheet(ws: Worksheet):
    columns = ['Language', 'Text']
    rows = [
        ['python', '-' * 10],
        ['java', 'j' * 10],
        ['c#', '*' * 5],
        ['c++', '0' * 50],
    ]

    for i, value in enumerate(columns, 1):
        ws.cell(row=1, column=i).value = value

    for i, row in enumerate(rows, 2):
        for j, value in enumerate(row, 1):
            cell = ws.cell(row=i, column=j)
            cell.value = value


wb = openpyxl.Workbook()
ws = wb.active

# Sheet 1
fill_sheet(ws)

# Sheet 2
ws = wb.create_sheet('set_column_size')
fill_sheet(ws)
set_column_size(ws)

wb.save('excel.xlsx')
