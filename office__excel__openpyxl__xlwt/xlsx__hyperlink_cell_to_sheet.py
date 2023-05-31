#!/usr/bin/env python3
# -*- coding: utf-8 -*-

__author__ = "ipetrash"


import openpyxl
from openpyxl.worksheet.hyperlink import Hyperlink


wb = openpyxl.Workbook()

sheet1 = wb.active

sheet2 = wb.create_sheet(title="Лист 2")

cell = sheet1.cell(row=1, column=1)
cell.hyperlink = f"#'{sheet2.title}'!K20"
cell.value = "Go"
cell.style = "Hyperlink"

cell = sheet1.cell(row=2, column=1)
cell.hyperlink = Hyperlink(
    location=f"'{sheet2.title}'!K20",
    ref="",
)
cell.value = "Go 2"
cell.style = "Hyperlink"

wb.save("excel.xlsx")
