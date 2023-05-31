#!/usr/bin/env python3
# -*- coding: utf-8 -*-

__author__ = "ipetrash"


import openpyxl


wb = openpyxl.Workbook()

ws = wb.active
ws.title = "10"
ws.sheet_view.zoomScale = 10

ws = wb.create_sheet("20")
ws.sheet_view.zoomScale = 20

ws = wb.create_sheet("50")
ws.sheet_view.zoomScale = 50

ws = wb.create_sheet("100")
# ws.sheet_view.zoomScale = 100

ws = wb.create_sheet("200")
ws.sheet_view.zoomScale = 200

ws = wb.create_sheet("500")
ws.sheet_view.zoomScale = 500

wb.save("excel.xlsx")
