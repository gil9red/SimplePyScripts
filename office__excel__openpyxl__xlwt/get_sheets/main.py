#!/usr/bin/env python3
# -*- coding: utf-8 -*-

__author__ = "ipetrash"


import openpyxl


wb = openpyxl.load_workbook("excel.xlsx")
print(wb.sheetnames)  # ['Sheet', 'auto', 'Students', 'Students1']

for name in wb.sheetnames:
    ws = wb[name]
    print(name, ws)
